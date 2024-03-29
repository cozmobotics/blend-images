# based on
# opencv tutorial --> https://opencv-python-tutroals.readthedocs.io/en/latest/py_tutorials/py_core/py_image_arithmetics/py_image_arithmetics.html

# todo:
# "image stabilisation" = zoom/pan/rotate similar images so they fit together 
# slowly pan panoramic images across screen 
# display information (height/width, date, filename,...) --> not so easy 
# graphical menu 
# play videos .... using os.system (mplayer filename)  
# react to mouse interaction
# invent some more fancy transitions
# change fade and duration at runtime
# make video-codec selectable by parameter
# try TAPI / OCL https://learnopencv.com/opencv-transparent-api/
# change parameters at runtime 
# on start, grab (capture) screen as first image and do some interesting transitions 
# sort images by date, aspect ratio,...
# print compile date, write date tothe \"done\"s
# when freezing picture, allow to move back and forth 
# mask with a growing circle --> better for change between landscape and portrait? 
# transition i: image has white margin (workaround: -bg #808080)
# copy path: Add to existing cliboard instead of replacing it (C .. add, c..replace)
# error when supplying an illegal value to -t parameter
# allow multiple keypresses (like c and p) to be evaluated with one picture 
# when playing from a list of pictures, the first picture is not blended in but starts immediately
# use mouse to switch to next picture 
# when scanning for files, write file list to disk. Write date of each subdir. On startup, compare subdir dates and re-scan newer subdirs
# in writeToPic, write the comment (if any)
# find startimage: make it faster (search for startimage when adding files to array, not when playing the show) 
# keep comments and display them 
# bug: after playing video with duration != -1, "missed" transitions are carried out --> bugfix: restart time  

# done: 
# subdirectries
# limit depth of subdirs
# stop "3rd picture" while blending = freeze
# copy image path+name to clipboard
# loop: do not read and discard the directory every time 
# opencv cannot read a filename containing funny characters 
# option: manually input height and width 
# error handling ... AttributeError: 'NoneType' object has no attribute 'shape'
# Error when blendig images with different color depths (ignore images with color depth other than 3)
# take only newest files. Paramater -a --age: age in days, -1 (default): all files
# convert to grayscale
# blend function completely rewritten --> better timing
# record slideshow as video (still experimental)
# "c" (copy file location) does not work corretly when moving backwards --> OK
# mask inverted
# first and last transition must be a fade
# syntax for masking effects, random effects
# transition masked: on some images, parts of old image remain --> workaround
# change background color when padding images
# hide mouse cursor --> not so easy --> workaround 
# fade time = 0 --> does not switch automatically .... why? --> works now. (Don't know why)
# function keys, cursor keys 
# 2021-05-15 light curve(s)
# 2021-05-17 save picture
# 2021-05-17 scripting 
# 2021-05-17 parameter "mask" renamed to "match"
# 2021-05-17 new parameter "notmatch"
# 2021-05-?? parameters --Match and --Nomatch (case sensitive)
# 2021-05-22 parameter --verbose
# 2021-05-22 bugfix for pause (pressing space/backspace does not terminate pause)
# 2021-05-22 on starup, print pwd and parameters
# 2021-05-25 fixed bug: when writing video, symbol w not found 
# 2021-05-29 transitions i (inverted) and x (xor)
# 2021-05-30 Output of builddate and -time (compiled version only)
# 2021-05-30 renamed output window from dst to blendPics
# 2021-05-30 parameter -t: value all added
# 2021-06-xx parameter "--limit": limit total time by randomly selecting a certain percentage 
# 2021-06-xx img1/img2/oddEven --> imgNew/imgOld
# 2021-06-xx key "e" invoke external command for current image
# 2021-07-18 looping back past first image 
# 2022-02-25 enable/disable screensaver 
# 2022-02-25 limit show to (approx.) time
# 2022-02-25 "blackout": leave screen black after last picture, wait for keypress
# 2022-10-30 plendPics.py is usable as a python mudule 
# 2022-10-30 Screensaver blendPicsScr.py imports plendPics.py 
# 2022-10-30 Screensaver-mode --scr (immediately stop with any key pressed (mouse not yet implemented))
# 2022-11-02 key "?" writes filename and -path on screen (redundant, Alt-Tab does the same)
# 2022-11-03 Screensaver-mode: quit on mouse activity
# 2022-11-08 bugfix: enable help in compiled version (see https://stackoverflow.com/questions/69007351/pyinstaller-is-not-compatible-with-python-argparse-h )
# 2022-12-04 do not enlarge small pictures (parameter -e), skip too small images
# 2022-12-04 Screensaver-mode: allow left and right arrow keys to change images 
# 2023-02-03 parameter --log, function logMessage() todo: replace all print commands with logMessage()
# 2023-02-13 when parameter -i is given: ignore comments, i.e. ignore all after a # sign (will cut a filename, if it contains a # sign)
# 2023-02-13 more significant error messages
# 2023-02-13 bugfix: enable mouse AFTER moving mouse pointer to 0,0 --> screensaver reacts to mouse 
# 2023-04-13 added complex sort options, added module exifRoutines
# 2023-04-15 added video playback (withe external player, default = mplayer)  
# 2023-04-19 added argument "startimage" 
# 2023-04-20 added ".mts" to list of extensionsVideo 
# 2023-04-26 bugfix when an input line contains more than one #'es 
# 2023-04-29 bugfix: F2, F3, F4,...  program crashes. NameError: name 'durationTime' is not defined
# 2023-05-02 load filenames from spreadsheet 
# 2023-05-03 only one command line parameter: interprete this parameter as path 
# 2023-05-05 bugfix: crash when filename in spreadscheet is empty (commented out)
# 2023-06-xx moved some functions to getPics.py
# 2023-07-09 more sorting options

import numpy as np
import cv2 as cv
import time
import os
import argparse
import re
import random
import clipboard
import pyautogui # for hiding cursor
import math # for sinus in light curve
import sys  # for printing arguments and enabling help in compiled version
from datetime import datetime # for logfile
import exifRoutines
import ctypes # to disable/enable screensaver (Windows)
import openpyxl # for spreadsheet 
from collections import namedtuple
import getPics

SPI_SETSCREENSAVEACTIVE = 17

#----------------------------------------------------------------
def logMessage (level, text):
	# text = '#' + text
	if level <= args.verbose:
		print (text)
	
	if (args.log !=  ""):
		try:
			logFile = open (args.log, 'a')
		except Exception:
			print ("--- unable to write to logfile " + args.log)
		else:
			logFile.write (text + '\n')
			logFile.close ()

#----------------------------------------------------------------
# https://bytes.com/topic/python/answers/24703-screensaver
def EnableScreenSaver(on):
	ctypes.windll.user32.SystemParametersInfoA(SPI_SETSCREENSAVEACTIVE, on, None, 0)

#----------------------------------------------------------------
# https://forum.opencv.org/t/filename-contains-character/3045/3
def imread_funny (filename):
	'''workaround because cv.imread() cannot handle non-ascii characters (funny characters)'''
	success = True
	image = None
	filename = filename.strip() # trailing newline causes a problem

	try: 
		f = open(filename, "rb")
		b = f.read()
		f.close()
	except Exception as e:
		# print ('>>>', e)
		success = False
		logMessage (0, '#>>>' + str(e))
	
	# 2023-02-13 more significant error messages
	
	try:
		b = np.frombuffer(b, dtype=np.int8)
	except Exception as e:
		# print ('>>>', e)
		success = False
		logMessage (0, '#>>>' + str(e))

	try:
		image = cv.imdecode(b, cv.IMREAD_COLOR);
	except Exception as e:
		# print ('>>>', e)
		success = False
		logMessage (0, '#>>>' + str(e))
	
	return (image, success)

#----------------------------------------------------------------
def scaleImage (tempImg, screenWidth, screenHeight, padColor, convertToGray):
	'''bring an image to the desired height and width, padding it to adjust for aspect ratio'''
	c = 3 # workaround 
	
	success = True
	
	try:
		(h,w,c) = tempImg.shape
	except Exception as e:
		# print (">>>> ERROR : ", str(e))
		logMessage (0, "#>>>> ERROR : " + str(e))
		success = False
	
	if (success):
		if (w>h):
			pl = "l"
		if (h>=w):
			pl = "p"
		if (not pl in args.portrait_landscape):
			success = False
			logMessage (2, "# portrait/landscape filter")
		
		if (w < screenWidth / 2) and (h < screenHeight / 2):
			success = False
			logMessage (2, "# image too small")


	if (success):
		aspectRatioImage = (w+1) / (h+1) 
		aspectRatioScreen = screenWidth / screenHeight
		
		if (w > screenWidth) or  (h > screenHeight) or (args.enlarge == "yes"):
		# if (w > screenWidth) or  (h > screenHeight) or (args.enlarge > 0):
			if (aspectRatioImage > aspectRatioScreen):
				newHeight = int(screenWidth / aspectRatioImage)
				newWidth = screenWidth
			else:
				newWidth = int(w * screenHeight / h)
				newHeight = screenHeight
			try:
				tempImg = cv.resize(tempImg, (newWidth, newHeight), cv.INTER_AREA)
			except Exception as e:
				logMessage (0, "#>>>" + str(e))
				success = False
		else:
			newWidth = w
			newHeight = h
			
		# print ("newWidth: ", newWidth, " newHeight:", newHeight)
		
		topBottom = int((screenHeight - newHeight) / 2)
		leftRight = int((screenWidth - newWidth)  / 2)
		color = [0, 0, 0]

		blackImg = np.zeros((screenHeight, screenWidth, 3), np.uint8)
		blackImg[:] = padColor
		blackImg[topBottom:topBottom + newHeight, leftRight:leftRight + newWidth] = tempImg
		tempImg = blackImg
	
	if (c != 3):
		if c == 1:
			tempImg = cv.cvtColor(tempImg, cv.COLOR_GRAY2BGR) # convert because we need a color depth of 3
		else:
			success = False
			logMessage (0, "#>>>> ERROR: image has color depth of " + str(c))
	
	if (convertToGray != 0):
		tempImg = cv.cvtColor(tempImg, cv.COLOR_BGR2GRAY)
		tempImg = cv.cvtColor(tempImg, cv.COLOR_GRAY2BGR) # convert back because we still need a color depth of 3 
	
	return (tempImg, success)
	
#----------------------------------------------------------------
def addInverted(img1,bright1,img2,bright2):
	
	imgA = img1
	imgB = img2
	brightA = bright1
	brightB = bright2

	# if (args.verbose >= 10):
		# print (int(100 * brightB), end = '\r')

	Negative = cv.bitwise_not(imgA) 
	gray = cv.cvtColor(imgB, cv.COLOR_BGR2GRAY)
	ret,mask = cv.threshold(gray,255 - int(255 * brightA) + 1,255,cv.THRESH_BINARY)

	if (ret):
		maskInverted = cv.bitwise_not (mask)
		maskedA = cv.bitwise_and (imgA, imgA, mask=mask)
		
		if (brightB > 0.5):
			Negative = cv.addWeighted(imgB, 2*(brightB -0.5), Negative, 2*brightA, 0)
		
		maskedB = cv.bitwise_and (Negative, Negative, mask=maskInverted)
		res = cv.bitwise_or (maskedA, maskedB)
		# cv.imshow ('debug', maskedA)
	else:
		res = cv.addWeighted(img1,bright1,img2,bright2,0)
	
	return res


	
	# if (args.verbose >= 10):
		# print (int(100 * brightB), end = '\r')

		# cv.imshow ('imgB',imgB)
		# cv.imshow ('imgA',Negative)
	
	return res

#----------------------------------------------------------------
def addXor(img1,bright1,img2,bright2):
	
	imgA = img1
	imgB = img2
	brightA = bright1
	brightB = bright2
	
	# if (args.verbose >= 10):
		# print (int(100 * brightB), end = '\r')

		# cv.imshow ('imgB',imgB)
		# cv.imshow ('imgA',imgA)

	subFactor = 400
	
	if (brightB < brightA):
		temp = cv.subtract (imgB, np.array([subFactor * (brightA - 0.5)]))
		res = cv.bitwise_xor (temp,imgA)
		# if (args.verbose >= 10):
				# cv.imshow ('B',temp)
				# cv.imshow ('A',imgA)
				# cv.waitKey(1)
	else:
		temp = cv.subtract (imgA, np.array([subFactor * (brightB - 0.5)]))
		res = cv.bitwise_xor (temp,imgB)
		# if (args.verbose >= 10):
			# cv.imshow ('B',imgB)
			# cv.imshow ('A',temp)
			# cv.waitKey(1)
	
	return res

#----------------------------------------------------------------
def addMasked(img1,bright1,img2,bright2, inverted, oldNew):
	'''add 2 images by calculating a mask which is derived from one of the images'''
	
	imgA = img1
	imgB = img2
	brightA = bright1
	brightB = bright2
	
	if (oldNew):
		gray = cv.cvtColor(imgA, cv.COLOR_BGR2GRAY)
		# b, g, r = cv.split(imgA)
		# gray = r
	else:
		gray = cv.cvtColor(imgB, cv.COLOR_BGR2GRAY)
		# b, g, r = cv.split(imgB)
		# gray = r

	
	if (inverted):
		ret,mask = cv.threshold(gray,255 - int(255 * brightB) + 1,255,cv.THRESH_BINARY)
		mask = cv.bitwise_not (mask)
	else:
		ret,mask = cv.threshold(gray,int(255 * brightB) + 1,255,cv.THRESH_BINARY)

	if (ret):
		maskInverted = cv.bitwise_not (mask)
		maskedA = cv.bitwise_and (imgA, imgA, mask=mask)
		maskedB = cv.bitwise_and (imgB, imgB, mask=maskInverted)
		res = cv.bitwise_or (maskedA, maskedB)
		# cv.imshow ('debug', maskedA)
	else:
		res = cv.addWeighted(img1,bright1,img2,bright2,0)
	
	return res
#----------------------------------------------------------------
def sinusFunc (x, percent):
	y1 = percent / 100 * ((math.sin((x - 0.5) * math.pi) * 1 + 1))
	y2 = (100 - percent)  * x
	y = (y1 + y2) / 2
	return y
	
#----------------------------------------------------------------
def savePic (imgToWrite):
	found = False
	count = 0
	
	pathToWrite = args.path
	testPath = os.path.join (args.path, 'blendPics')
	if os.path.exists (testPath):
		if (os.path.isdir(testPath)):
			pathToWrite = testPath
	
	while (not found):
		filenameToWrite = 'blendPics' + str (count).zfill(4) + '.jpg'
		filenameToWrite = os.path.join (pathToWrite, filenameToWrite)
		if (os.path.exists(filenameToWrite)):
			count = count + 1
		else:
			found = True
	
	# if (args.verbose >= 1):
		# print ("writing image to ", filenameToWrite)
	logMessage (1, "# writing image to " + filenameToWrite)
	cv.imwrite (filenameToWrite, imgToWrite)
		
#----------------------------------------------------------------
def evalKey (key):
	global convertToGray
	global durationTime
	global fadeTime
	global transDict
	global all
	
	if (key > 0):
		if (key == ord('g')):
			convertToGray = abs(convertToGray - 1)
			# if (args.verbose >= 1):
				# print ("convertToGray:", convertToGray)
			key = -1
		elif (key == KEY_F2):
			durationTime = durationTime / 1.5
			if (args.verbose >= 1):
				print ("#duration: ", "%.2f" % durationTime)
		elif (key == KEY_F3):
			if (durationTime < 0.1):
				durationTime = 0.5
			else:
				durationTime = durationTime * 1.5
			if (args.verbose >= 1):
				print ("#duration: ", "%.2f" % durationTime)
		elif (key == KEY_F4):
			fadeTime = fadeTime / 1.5
			if (args.verbose >= 1):
				print ("fade: ", "%.2f" % fadeTime, end='\r' )
		elif (key == KEY_F5):
			if (fadeTime < 0.1):
				fadeTime = 0.5
			else:
				fadeTime = fadeTime * 1.5
			if (args.verbose >= 1):
				print ("#fade: ", "%.2f" % fadeTime)
		# elif (key < 128 and chr(key) in 'bondlix'):           # < 128 : avoid error when pressing function/cursor keys
		elif (key < 128 and chr(key) in all):           # < 128 : avoid error when pressing function/cursor keys
			transDict = setTransitions (chr(key), transDict)
			if (args.verbose >= 1):
				print ("#Transitions: ", listTransitions(transDict))
			# print (chr(key), transDict) # debug
		else:
			return key
	
	return -1

#----------------------------------------------------------------
def writeToPic(imageToWriteTo, textToWrite, lineNr = 0 ):
	global writeToPicLineNr
	
	if lineNr == -1:
		lineNr = writeToPicLineNr
		writeToPicLineNr = writeToPicLineNr + 1
	else:
		writeToPicLineNr = lineNr + 1
		
	lineHeight = 50
	new_image = cv.putText(
	img = imageToWriteTo,
	text = textToWrite,
	org = (200, 200 + lineNr * lineHeight),
	fontFace = cv.FONT_HERSHEY_DUPLEX,
	fontScale = 1.0,
	color = (125, 246, 55),
	thickness = 1
	)
	
	cv.imshow ('blendPics', new_image)
	cv.waitKey (1)


#----------------------------------------------------------------

def blendWrapper (img1, img2, fadeTime, duration, transition, description, isVideo1, isVideo2, filename1, filename2):
	global black
	global args
	if not isVideo1 and not isVideo2:
		return blend (img1, img2, fadeTime, duration, transition, description)
	
	elif isVideo2:
		if args.videoplayer != "none":
			img2 = black
			cv.imshow ("blendPics", img2) 
			cv.waitKey (1)
			# os.system ("mplayer " + filename2)
			os.system (args.videoplayer + " " + filename2)
		return 0

	elif isVideo1:
		img1 = black
		return blend (img1, img2, fadeTime, duration, 'b', description)

#----------------------------------------------------------------

def blend (img1, img2, fadeTime, duration, transition, description):
		'''do the actual transition from one image to the other and evaluate users keypresses'''
		global transDict
		global rawTransDict
		global pause
		global debugFrameCounter
		global videoInterval
		global out
		global args

		global KEY_UP   
		global KEY_DOWN 
		global KEY_LEFT 
		global KEY_RIGHT
		global KEY_F1   
		global KEY_F2   
		global KEY_F3   
		global KEY_F4   
		global KEY_F5   
		global KEY_F6   
		global KEY_F7   
		global KEY_F8   
		global KEY_F9   
		global KEY_F11  
		global KEY_F12  
		global KEY_F1   
		global KEY_PGUP 
		global KEY_PGDN 
		global KEY_HOME 
		global KEY_END  

		if (transition == ""):
			possible = []
			if (transDict['b']): 
				possible.append ({'b': 1,'o': 0,'n': 0,'d': 0,'l': 0,'i':0,'x': 0})
			if (transDict['i']): 
				possible.append ({'b': 0,'o': 0,'n': 0,'d': 0,'l': 0,'i':1,'x': 1})
			if (transDict['x']): 
				possible.append ({'b': 0,'o': 0,'n': 0,'d': 0,'l': 0,'i':0,'x': 1})
			if (transDict['o']): 
				if (transDict['d']):
					possible.append ({'b': 0,'o': 1,'n': 0,'d': 1,'l': 0,'i':0,'x': 0})
				if (transDict['l']):
					possible.append ({'b': 0,'o': 1,'n': 0,'d': 0,'l': 1,'i':0,'x': 0})
			if (transDict['n']): 
				if (transDict['d']):
					possible.append ({'b': 0,'o': 0,'n': 1,'d': 1,'l': 0,'i':0,'x': 0})
				if (transDict['l']):
					possible.append ({'b': 0,'o': 0,'n': 1,'d': 0,'l': 1,'i':0,'x': 0})
			
			# if nothing is possible, we do a fade
			if (len (possible) == 0):
				possible.append ({'b': 1,'o': 0,'n': 0,'d': 0,'l': 0,'x': 0})

			numPossible = len (possible)
			transIndex = random.randrange(0, numPossible)
			tempTransDict = possible[transIndex]

		else:
			tempTransDict =  setTransitions(transition, rawTransDict)
			
				
		if (args.verbose >= 2):
			print ("*t|" + listTransitions(tempTransDict))
		
		# print (tempTransDict)
		
		timeNextFrame = time.time() + videoInterval
		
		bright1 = 1
		bright2 = 0


		
		dst = img1

		quit = False
		freeze = False
		timeStart = time.time()
		timeEnd = timeStart + fadeTime
		timeBefore = 0
		timeAfter = 0
		key = -1
		
		while (time.time() < timeEnd):
			
			if (freeze):
				timeStart = time.time() - timeBefore
				timeEnd   = time.time() + timeAfter
			else:
				bright = (time.time() - timeStart) / (timeEnd - timeStart)
				bright = sinusFunc (bright, 100)
				bright1 = 1 - bright
				bright2 = bright
			

			if (tempTransDict['b']):
				dst = cv.addWeighted(img1,bright1,img2,bright2,0)
			elif (tempTransDict['i']):
				dst = addInverted(img1,bright1,img2,bright2)
			elif (tempTransDict['x']):
				dst = addXor(img1,bright1,img2,bright2)
			else:
				dst = addMasked(img1,bright1,img2,bright2,tempTransDict['l'],tempTransDict['o'])
			
			cv.imshow('blendPics',dst)
			if ((args.output != '') and (time.time() > timeNextFrame)):
				out.write(dst)
				timeNextFrame = timeNextFrame + videoInterval
				debugFrameCounter = debugFrameCounter + 1

			tempKey = cv.waitKeyEx(1) 
			if (tempKey != -1):
				key = tempKey
				
				if (args.scr == "yes") and not (tempKey in [KEY_LEFT, KEY_RIGHT]):	# when blendPics acts as screensaver, quit immediately
					# if (args.verbose > 1):
						# print ("screensaver mode, stopping immediately")
					logMessage (1, "# screensaver mode, stopping immediately")
					sys.exit()	
				
				if (tempKey == ord ('f')):
					freeze = not freeze
					if (args.verbose >= 1):
						print ('frozen ',freeze, '   ',end = '\r') # +++ python3 only 
					if (freeze):
						timeBefore = time.time() - timeStart
						timeAfter = timeEnd - time.time()
				elif ((tempKey == KEY_UP) or (tempKey == KEY_DOWN)):
					if (freeze):
						if (tempKey == KEY_UP):
							timeBefore = timeBefore + fadeTime / 10
							timeAfter  = timeAfter + fadeTime / 10
						else:
							timeBefore = timeBefore - fadeTime / 10
							timeAfter  = timeAfter - fadeTime / 10
					tempKey = -1
				# elif (tempKey < 128 and chr(tempKey) in 'bondlix'):
				elif (tempKey < 128 and chr(tempKey) in all):
					transDict = setTransitions (chr(tempKey), transDict)
					if (args.verbose >= 1):
						print ("Transitions: ", listTransitions(transDict))
					# print (chr(tempKey), transDict) # debug
					tempKey = -1
				elif (tempKey == ord('s')):
					savePic (dst)
					tempKey = -1
				else:
					key = evalKey(tempKey)
					
					
		
		# blending finished
		
		#workaround for some pictures/transitions when parts of the old image remain
		dst = img2
		cv.imshow ('blendPics',dst)
		cv.waitKey (1)
		
		# if ((key == 27) or (key == ord('q')) or (key == ord(' ')) or (key == 8)):
		if (key in (27,ord('q'),32,8,KEY_LEFT,KEY_RIGHT)):
			quit = True

		timeThen = time.time() + duration
		while (pause or (time.time() < timeThen) or (duration < 0)) and (not quit):

			if ((args.output != '') and (time.time() > timeNextFrame)):
				out.write(dst)
				timeNextFrame = timeNextFrame + videoInterval
				debugFrameCounter = debugFrameCounter + 1
			
			tempKey = cv.waitKeyEx(1)
			
			if ((tempKey != -1) and(args.scr == "yes") and not (tempKey in [KEY_LEFT, KEY_RIGHT])):	# when blendPics acts as screensaver, quit immediately
				# if (args.verbose > 1):
					# print ("screensaver mode, stopping immediately")
				logMessage (1, "# screensaver mode, stopping immediately")
				sys.exit()	

			
			tempKey = evalKey (tempKey)
			
			if (tempKey == ord('p')):
				pause = not pause
				if (args.verbose >= 1):
					print ('#pause ',pause, '   ') 
			elif (tempKey != -1):
				key = tempKey
				if ((key == ord(' ')) or (key == 8) or (key == KEY_LEFT) or (key == KEY_RIGHT)):
					break
				if ((key == 27) or (key == ord('q'))):
					quit = True
				if (tempKey == ord('?')):
					writeToPic (img2, description)
		# print ("key:", key)
		return (key)
		
		# --- blend ---

#----------------------------------------------------------------
def setTransitions (transitionsDef, oldDict):
	'''read a string like bondlix and switch entries in transitionsDef according to the letters in the string'''
	# global transDict
	newDict = oldDict.copy()
	
	for character in transitionsDef:
		newDict[character] = not newDict[character]
	
	return (newDict)

#----------------------------------------------------------------
def listTransitions (dict):
	'''create a string to output the transitions available'''
	
	res = ''
	keys = dict.keys()
	
	for key in keys:
		if (dict[key]):
			res = res + key
			
	return res

#----------------------------------------------------------------
def maskingAllowed (dict):
	'''see if we are allowed to use masking'''
	
	return (dict['o'] or dict['n']) and (dict['d'] or dict['l'])
#----------------------------------------------------------------
def readBgColor (inString):
	'''read a string like #aabbcc and convert it to BGR-value (not RGB!)'''
	if (inString[0] != '#') or (len(inString) != 7):
		return ([0,0,0])
	
	inString = inString[1:7]
	
	res = []
	
	# for countDoubles in range (3):
	for countDoubles in range (2,-1,-1):
		hex = 0
		substr = inString[2 * countDoubles : 2 * countDoubles + 2]
		# print (substr)
		hex = int (substr, 16)
		# print ('dez:',hex)
		res.append (hex)
		
	return res
#---------------------------------------------------------------------------
def fromSpreadsheet (spreadSheeetName):
	
	files = []
	
	try:
		wb = openpyxl.load_workbook(filename=spreadSheeetName)
		ws = wb.worksheets[0]
	except Error as e:
		print (str(e))
		return ([])
	
	for lineCount in range (1, ws.max_row + 1):
		line = ""
		filename = ws["B" + str(lineCount)].value 
		try:
			filename = filename.strip()
		except Exception:
			filename = ""
		comment = ws["C" + str(lineCount)].value
		
		if (filename != None): 
			files.append(filename)

	return files

#----------------------------------------------------------------
def mouseEvent (event,x,y,flags,param):
	global args
	global mouseCount
	global mouseOldX
	global mouseOldY
	
	if mouseCount == 0:
		mouseOldX = x
		mouseOldY = y
	else:
		if (args.scr == "yes"): 
			if (abs (mouseOldX - x) > 10) or (abs (mouseOldY - y) > 10):
				# if (args.verbose > 1):
					# print ("screensaver mode, mouse moved, stopping immediately")
				logMessage (0, "# screensaver mode, mouse moved, stopping immediately")
				sys.exit()	
		
	mouseCount = mouseCount + 1
	
	if (args.verbose > 1):
		print (event,x,y,flags,param)
	
	if (args.scr == "yes") and (event != 0):	# when blendPics acts as screensaver, quit immediately
		# if (args.verbose > 1):
			# print ("screensaver mode, mouse clicked, stopping immediately")
		logMessage (0, "# screensaver mode, mouse clicked, stopping immediately")
		sys.exit()	

#----------------------------------------------------------------
def doIt (argumentsDoIt):
	global videoInterval
	global pause
	global transDict
	global rawTransDict
	global all
	global args
	global convertToGray
	
	global durationTime
	global fadeTime
	
	global mouseCount
	global mouseOldX
	global mouseOldY
	global writeToPicLineNr
	global welcome
	
	writeToPicLineNr = 0

	
	args = argumentsDoIt

	rawTransDict = {
		'b': False,			# fade --> blend
		'o': False,		# maskOld
		'n': False,		#maskNew
		'd': False,	# dark
		'l': False,	# light
		'i': False,	# invert
		'x': False	# xor
	}

	global KEY_UP   
	global KEY_DOWN 
	global KEY_LEFT 
	global KEY_RIGHT
	global KEY_F1   
	global KEY_F2   
	global KEY_F3   
	global KEY_F4   
	global KEY_F5   
	global KEY_F6   
	global KEY_F7   
	global KEY_F8   
	global KEY_F9   
	global KEY_F11  
	global KEY_F12  
	global KEY_F1   
	global KEY_PGUP 
	global KEY_PGDN 
	global KEY_HOME 
	global KEY_END  
	global black
	global startImageFound

	KEY_UP   = 0x260000
	KEY_DOWN = 0x280000
	KEY_LEFT = 0x250000
	KEY_RIGHT= 0x270000
	KEY_F1   = 0x700000
	KEY_F2   = 0x710000
	KEY_F3   = 0x720000
	KEY_F4   = 0x730000
	KEY_F5   = 0x740000
	KEY_F6   = 0x750000
	KEY_F7   = 0x760000
	KEY_F8   = 0x770000
	KEY_F9   = 0x780000
	KEY_F11  = 0x790000
	KEY_F12  = 0x7a0000
	KEY_F1   = 0x7b0000
	KEY_PGUP = 0x210000
	KEY_PGDN = 0x220000
	KEY_HOME = 0x240000
	KEY_END  = 0x230000
	
	mouseCount = 0
	mouseOldX = 0
	mouseOldY = 0
	
	
	
	

	extensionsPhoto = ('.png', '.jpg', '.jpeg', '.jfif', '.tiff', '.bmp' , '.webp', '.gif') 
	extensionsVideo = ('.mp4', '.mpeg', '.avi', '.mov','.mts') 


	key = 0
	pause = False

	logMessage (100, "# ---------------------------------------------------")
	now = datetime.now()
	logMessage (100, "# " + now.strftime("%Y-%m-%d, %H:%M:%S"))
	
	if (args.verbose >= 1):
		print ("# blendPics by Martin Piehslinger")
		print ("# *******************************")
		print ("")

		try:
			from builddate import builddateString
			print ("Version from ", builddateString)
		except ImportError:
			pass

		print ('# current working directory: ', os.getcwd(), ', command line parameters:', args)

	if ("Y" in args.screensaver.upper()):
		EnableScreenSaver(0)

	all = "bondlix"
	if (args.transition == "all"):
		transitionString = all
	elif (args.transition == "fade"):
		transitionString = "b"
	elif (args.transition == "mask"):
		transitionString = "ol"
	elif (args.transition == "imask"):
		transitionString = "od"
	else: 
		transitionString = args.transition
	transDict = setTransitions(transitionString, rawTransDict)

	if (args.verbose >= 1):
		print ("Transitions: ", listTransitions(transDict))

	if maskingAllowed (transDict):
		blendString = 'nd'
	else:
		blendString = 'b'

	fadeTime = args.fade
	durationTime = args.duration
	convertToGray = args.gray

	filenamesAll = []

	if (args.input == ''):
		# if (args.verbose >= 1):
			# print ("searching images on " + args.path)
		logMessage (1, "# searching images on " + args.path)
		fileList = getPics.getListOfFiles(args.path, args.subdirs)

		for filename in fileList:
			addFile = True
			if ((filename.lower().endswith(extensionsPhoto)) or (filename.lower().endswith(extensionsVideo))):
				if (not re.search(args.match,filename,flags=re.IGNORECASE)):
					addFile = False
				if (not re.search(args.Match,filename)):
					addFile = False
				if (args.notmatch != '' and re.search(args.notmatch,filename,flags=re.IGNORECASE)):
					addFile = False
				if (args.NotMatch != '' and re.search(args.NotMatch,filename)):
					addFile = False
				if (args.age > 0):
					age = time.time() - os.path.getmtime(filename) ### +++ aus exif nehmen
					age = age / 3600 / 24 # age in days
					# print ("age: ",age," max age:" , args.age, addFile, age > args.age)
					if (age > args.age):
						addFile = False
				if (addFile):
					filenamesAll.append (filename)
				
				# trying some more sophisticated evaluation, especially when doing a negative lookahead assertion
				# does not work....
				# result = re.search(args.match,filename)
				# if (result):
					# (start,end) = result.span()
					# print (start,end,filename)
					# if (end > start):
						# filenames.append (os.path.join(args.path, filename))


			
		# if (args.random == 0):
			# print ("sorting...")
			# filenames.sort()
			
		# print (filenames)
	else:
		if args.input.lower().endswith('.xlsx'):
			filenamesAll = fromSpreadsheet (args.input)
		else: 
			try:
				inputList = open (args.input,'r')
			except Exception as e:
				print (e)
				sys.exit() 
			filenames = inputList.readlines()
			
			for filename in filenames:
				# 2023-02-13
				if '#' in filename:
					(a,b) = filename.split('#', 1)
					filename = a
				filename = filename.strip()  # to avoid blank lines on the screen 
				filenamesAll.append (filename)
			# filenamesAll = filenames
			inputList.close()

	NumFiles = len(filenamesAll)
	LenShow = int (NumFiles * (fadeTime + durationTime) / 60)

	if ((args.limit > 0) and (args.limit < LenShow)):
		probability = 100 * args.limit / LenShow  # +++ inaccurate! Take fade time into account!
	else:
		probability = 100
	# print ("probability ", probability)

	# if (args.verbose >= 1):
		# print (NumFiles, "files found", " total duration approx", LenShow, " minutes,  limiting to ", args.limit, " minutes. Probability=", int(probability), "%")
	logMessage (1, "# " + str(NumFiles) + " files found," + " total duration approx " + str(LenShow) + " minutes, limiting to " + str(args.limit) + " minutes, Probability=" + str(int(probability)) + "%")
	if (NumFiles == 0):
		sys.exit()

	if ((args.width > 0) and (args.height > 0)):
		flag = cv.WND_PROP_AUTOSIZE
		mode = cv.WINDOW_AUTOSIZE
		w = args.width
		h = args.height
	else:
		flag = cv.WND_PROP_FULLSCREEN
		mode = cv.WINDOW_FULLSCREEN
		w = -1
		h = -1
		


	# Window in full-screen-mode, determine aspect ratio of screen
	cv.namedWindow('blendPics', flag)
	cv.setWindowProperty('blendPics',flag,mode)
	if ((w < 0) or (h < 0)):
		(a,b,w,h) = cv.getWindowImageRect('blendPics')
	aspectRatioScreen =  (w+1) / (h+1)
	# if (args.verbose >= 1):
		# print ('Screen: ', w+1, h+1, "%.2f" % aspectRatioScreen)
	logMessage (1, '# Screen: ' +  str(w+1) + "x" + str(h+1) + ", " + str(aspectRatioScreen))


	if (args.output != ''):
		fourcc = cv.VideoWriter_fourcc(*'XVID')
		out = cv.VideoWriter(args.output, fourcc, args.fps, (w+1,  h+1))
	videoInterval = 1 / args.fps
	debugFrameCounter = 0
	timeBeginVideo = time.time()

	bgColor = readBgColor(args.background)

	black = np.zeros((h+1, w+1, 3), np.uint8)
	black[:] = bgColor
	
	img1 = black
	img2 = black
	cv.imshow('blendPics',black)
	# cv.waitKey(1000)
	
	
	welcome = np.zeros((h+1, w+1, 3), np.uint8)
	welcome[:] = bgColor
	writeToPic (welcome, 'Welcome to blendPics')
	

	FirstTime = True

	pyautogui.FAILSAFE = False
	pyautogui.moveTo(0,0) # to hide the mouse cursor
	cv.setMouseCallback('blendPics',mouseEvent) # 2023-02-13 enable mouse AFTER moving mouse pointer to 0,0

	loopCount = 0
	
	if (args.startimage != ""):
		writeToPic (welcome, "searching for " + args.startimage, lineNr = -1)

	while ((loopCount < args.loop) or (args.loop == -1)):

		writeToPic (welcome, "sorting...", -1)
		logMessage (1, "# sorting....")
		filenames = getPics.bubbleShuffle (filenamesAll, args.random, probability, args.sort)
		
		# if (args.random != 0):
			# print ("shuffle....")
			# random.shuffle (filenames) 


		IndexFiles = 0
		running = True
		Back = False
		isVideo1 = False
		isVideo2 = False

		while ((IndexFiles >= 0) and (IndexFiles < len(filenames)) and running):
			
			# script gives detailed specifications for the image (syntax: imgName|fadeTime|durationTime|transitionString) 
			if (filenames[IndexFiles].startswith('*')):
				commands = filenames[IndexFiles].split ('*')
				for command in commands:
					if (len(command) > 0):
						# print (command)
						(key,val) = command.split ('|')
						if (key == 'n'):
							imgName = val
						elif (key == 'f'):
							fadeTime = float(val)
						elif (key == 'd'):
							durationTime = float(val)
						elif (key == 't'):
							transitionString = val
							
			else:
				imgName = filenames[IndexFiles]
			
			if (args.verbose >= 2):
				# print ('*n|' + imgName + '*d|' + "%.2f" % durationTime + '*f|' + "%.2f" % fadeTime,end = '')
				logMessage (0, '*n|' + imgName + '*d|' + "%.2f" % durationTime + '*f|' + "%.2f" % fadeTime + " " + imgName)
			elif (args.verbose == 1):
				# print (imgName)
				logMessage (0, imgName)
				# logMessage (0, imgName + " # >" + exifRoutines.getImageTime(imgName) + "<") ###+++ debug
			
			if (imgName.lower().endswith(extensionsPhoto)):
				isVideo2 = False
				(img2, success) = imread_funny(imgName)
				if (success):
					(img2, success) = scaleImage(img2, w+1, h+1, bgColor, convertToGray)
			elif (imgName.lower().endswith(extensionsVideo)):
				isVideo2 = True
				success = True
			
			if (args.startimage != ""):
				# if imgName.endswith (args.startimage):
				if args.startimage in imgName:
					startImageFound = True
				success = startImageFound

			# this is redundant, but I want it to work also when using an input file
			if (not re.search(args.match,imgName,flags=re.IGNORECASE)):
				success = False
			if (not re.search(args.Match,imgName)):
				success = False
			if (args.notmatch != '' and re.search(args.notmatch,imgName,flags=re.IGNORECASE)):
				success = False
			if (args.NotMatch != '' and re.search(args.NotMatch,imgName)):
				success = False
			if (args.age > 0):
				age = time.time() - os.path.getmtime(imgName) ### +++ aus exif nehmen
				age = age / 3600 / 24 # age in days
				# print ("age: ",age," max age:" , args.age, success, age > args.age)
				if (age > args.age):
					success = False







			if (success):
				if (FirstTime):
					key = blendWrapper (img1, img2, fadeTime, durationTime, blendString, imgName, isVideo1, isVideo2, "", imgName) #{'b': 0,'o': 0,'n': 1,'d': 1,'l': 0})
				else:
					key = blendWrapper (img1, img2, fadeTime, durationTime, '', imgName, isVideo1, isVideo2, "", imgName)
				FirstTime = False
				Back = False
				if (key > 0):
					# print (chr(key))
					if ((key == ord('q')) or (key == 27)):
						if (args.verbose >= 1):
							print ("# interrupted by user")
						running = False

					elif ((key == 8) or (key == KEY_LEFT)):
						Back = True
						key = -1
					elif (key == ord('c')):
						if (args.verbose >= 1):
							print ("copying ", imgName, " to clipboard")
						clipboard.copy (imgName)
						key = -1
					elif (key == ord('C')):
						if (args.verbose >= 1):
							print ("copying ", imgName, " to clipboard")
						clipboard.copy (clipboard.paste() + '\r\n' + imgName)
						key = -1
					elif (key == ord('e')):
						os.system("start cmd /k \"" + imgName + '\"') # windows-specific! 

				
				img1 = img2
				isVideo1 = isVideo2
			else:
				# print ("Skipped ", imgName)
				logMessage (0, "# Skipped " + imgName)
				
				
			if (Back):
				IndexFiles = IndexFiles - 1
				if (IndexFiles < 0):
					IndexFiles = len(filenames) - 1
					loopCount = loopCount - 1
			else:
				IndexFiles = IndexFiles + 1


		loopCount = loopCount + 1

		if ((key == ord('q')) or (key == 27)):
			# if (args.verbose >= 1):
				# print ("break")
			logMessage (0, "# break")
			break
			
	# done, fade out
	img2 = black
	isVideo2 = False
	if maskingAllowed (transDict):
		blendString = 'od'
	else:
		blendString = 'b'


	blendWrapper (img1, img2, fadeTime, 0, blendString, imgName, isVideo1, isVideo2, "", "" ) # {'b': 0,'o': 1,'n': 0,'d': 1,'l': 0})
	cv.waitKey (1000)

	# if (args.verbose >= 1):
		# print ("Done.")
	logMessage (0, "# Done.")
		
	if ("Y" in args.blackout.upper()):
		print ('# In main window, press any key to exit')
		writeToPic (img2, 'Press any key to exit')
		cv.waitKey(0)
		
	cv.destroyAllWindows()
	if (args.output != ''):
		out.release()
		print (debugFrameCounter, 'frames written in ', time.time() - timeBeginVideo,' = ', debugFrameCounter/(time.time() - timeBeginVideo), ' fps'	)
	
	if ("Y" in args.screensaver.upper()):
		EnableScreenSaver(1)

#----------------------------------------------------------------

if __name__ == "__main__":

	if (len(sys.argv) == 2) and not (sys.argv[1].startswith('-')):		# one command line argument given
		args = namedtuple ( 
		"arguments", 
		"path subdirs duration fade transition all match Match notmatch NotMatch portrait_landscape  enlargeloop limit age gray random input verbose output width height fps background blackout screensaver scr log sort videoplayer startimage"
		)

		args.path = sys.argv[1] # this one and only parameter is the path
		
		args.subdirs = -1
		args.duration = 5
		args.fade = 1.5
		args.transition = "all"
		args.match = "."
		args.Match = "."
		args.notmatch = ""
		args.NotMatch = ""
		args.portrait_landscape  = "pl"
		args.enlarge = "no"
		args.loop = 1
		args.limit = -1
		args.age = -1.0
		args.gray = 0
		args.random = 0
		args.input = ""
		args.verbose = 1
		args.output = ""
		args.width = -1
		args.height = -1
		args.fps = 30.0
		args.background = "#000000"
		args.blackout = "yes"
		args.screensaver = "yes"
		args.scr = "no"
		args.log = ""
		args.sort = ""
		args.videoplayer = ""
		args.startimage = ""

	else:
		parser = argparse.ArgumentParser(description = "display all images in folder with nice transitions", epilog = "Esc/q=quit, p=pause on/off, f=freeze on/off, s(while blending)=save picture, c=copy filename to clipboard, b/o/n/d/l=change transition, left arrow or backspace=previous, right arrow or space=next, F2,F3,F4,F5=decrease/increase fade-time/duration" )
		parser.add_argument("-p", "--path", type=str, default=".", help="path where images are found")
		parser.add_argument("-s", "--subdirs", type=int, default=0, help="depth of subdirectories.  0 (default): no subdirs, -1: all subdirs")
		parser.add_argument("-d", "--duration", type=float, default="5", help="time image is shown [seconds]. -1 for manual switching")
		parser.add_argument("-f", "--fade", type=float, default="1.5", help="time of fading effect [seconds]")
		parser.add_argument("-t", "--transition", type=str, default="all", help="types of transition. Combination of the letters b,o,n,d,l,i,x")
		parser.add_argument("-m", "--match", type=str, default=".", help="mask filename (regex syntax, case insensitive)")
		parser.add_argument("-M", "--Match", type=str, default=".", help="mask filename (regex syntax, case sensitive)")
		parser.add_argument("-n", "--notmatch", type=str, default="", help="negative mask filename (regex syntax, case insensitive)")
		parser.add_argument("-N", "--NotMatch", type=str, default="", help="negative mask filename (regex syntax, case sensitive)")
		parser.add_argument("-pl", "--portrait_landscape", type=str, default="pl", help="filter portrait or landscape. p = portrait, l = landscape, pl (default) = both")
		parser.add_argument("-e", "--enlarge", type=str, default="no", help="enlarge small images, default = no ")
		parser.add_argument("-l", "--loop", type=int, default="1", help="nr. of loops, -1 = loop forever, default=1")
		parser.add_argument("-L", "--limit", type=int, default="-1", help="limit length of show to minutes (approx.) by randomly skipping images and maintaining the order ")
		parser.add_argument("-a", "--age",  type=float, default="-1.0", help="maximal age of file in days. -1.0 (default): all files")
		parser.add_argument("-g", "--gray", type=int, default="0", help="0 (default): color, all else: convert to grayscale")
		parser.add_argument("-r", "--random", type=int, default="-1", help="random shuffle. -1 (default): leave as is = depth-first-search, 0: sorted, 0..100: shuffle")
		parser.add_argument("-i", "--input", type=str, default="", help="input file, containing filenames [and parameters] (script)")
		parser.add_argument("-v", "--verbose", type=int, default="1", help="verbose ... 0=only errors, 1(Default)=print filenames, higher numbers = more information")
		parser.add_argument("-o", "--output", type=str, default="", help="output video file (very experimental)")
		parser.add_argument("-w", "--width", type=int, default="-1", help="width. -1 (default): automatic")
		parser.add_argument("-hh", "--height", type=int, default="-1", help="height. -1 (default): automatic")
		parser.add_argument("--fps", type=float, default=30.0, help="fps of output video (very experimental)")
		parser.add_argument("-bg", "--background", type=str, default="#000000", help="background color in hex values (rgb) like #0fcc80")
		parser.add_argument("-b", "--blackout", type=str, default="yes", help="yes/no. At the end of the show, keep the window dark until a key is pressed. Default = yes")
		parser.add_argument("-ss", "--screensaver", type=str, default="yes", help="yes/no. Disable screensaver. Default = yes")
		parser.add_argument("--scr", type=str, default="no", help="blendPics acts as screensaver (quit when a key is pressed), Default=no")
		parser.add_argument("--log", type=str, default="", help="filename to log actions for debug purposes. Empty string(default): do not log")
		parser.add_argument ("--sort", type=str, default="", help="sorting options ... syntax: tbd")
		parser.add_argument ("--videoplayer", type=str, default="", help="external command to play video. Default: system settings")
		parser.add_argument ("--startimage", type=str, default="", help="image name to start with (if not starting with first image)")

		args = parser.parse_args(sys.argv[1:]) # sys.argv[1:] for enabling help in compiled version

	black = None
	startImageFound = False
	doIt (args)

